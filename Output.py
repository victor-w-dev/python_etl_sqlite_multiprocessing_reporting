import calendar
import os
import numpy as np
import openpyxl
import pandas as pd
import xlsxwriter
from BSO.time_analysis import time_decorator
from win32com.client import Dispatch
import copy
#import threading


class ExcelOutput():
    """docstring forExcelOutput."""

    def __init__(self, countryname, periods_required, toprank, trades_general_dict, trades_byproduct_dict, report_type, currency='HKD',money='MN'):
        self.countryname = countryname
        self.trades_general_dict = copy.deepcopy(trades_general_dict)
        self.trades_byproduct_dict = copy.deepcopy(trades_byproduct_dict)
        self.report_type = report_type
        self._currency = currency
        self._money = money
        self._lastperiod = periods_required[-1]
        self._periods = periods_required
        self._excel_name = f"{(self.countryname).replace('/',',')}_{self._lastperiod}_{self._currency}_{self._money}.xlsx"
        self._toprank = toprank

    def money_conversion(self):
        dollar = {'HKD':1, 'USD':7.8}
        unit = {'TH':10**3, 'MN':10**6}

        self.trades_general_dict['figures'] = \
        self.trades_general_dict['figures'] / (dollar[self._currency]*unit[self._money])

        for k, v in self.trades_byproduct_dict.items():
            v['figures'] = v['figures']/(dollar[self._currency]*unit[self._money])

    def create_and_change_path(self):
        file_path="Output"+"/"+"R1_"+str(self._lastperiod)+"/"+self.report_type+"/"+self._currency+"/"+self._money
        if not os.path.exists(file_path):
            os.makedirs(file_path)
        self._file_path = file_path

    #@time_decorator
    def denotesymbol(data, datatype):
        if datatype=='figures':
            data[(abs(data)<0.5) & (abs(data)>0)] ='*'
            data.replace([np.nan, 0], '-',inplace=True)

        elif datatype=='share':
            data[(data<0.05) & (data>0)] ='*'
            data.replace([np.inf, -np.inf], '∞',inplace=True)
            data.replace([np.nan, 0], '-',inplace=True)

        elif datatype=='percent_change':
            data[((data<0.05) & (data>0))|((data>-0.05) & (data<0))] = 0.001
            data[((data>1000)&(data<np.inf))|((data<-1000)&(data>-np.inf))] = '..'
            data.replace([np.inf, -np.inf], '∞',inplace=True)
            data.replace([np.nan, 0], '-',inplace=True)
            data.replace(0.001, '*',inplace=True)
        return data

    #@time_decorator
    def export_results(self):
        self.money_conversion()

        self.create_and_change_path()
        writer = pd.ExcelWriter(f"{self._file_path}/{self._excel_name}", engine='xlsxwriter')

        ExcelOutput.denotesymbol(self.trades_general_dict['figures'], 'figures').to_excel(writer,sheet_name=f"{self._currency}_{self._money}",index=False,startrow=5, startcol=3,header=False)
        ExcelOutput.denotesymbol(self.trades_general_dict['percent_change'], 'percent_change').to_excel(writer,sheet_name=f"{self._currency}_{self._money}",index=False,startrow=5, startcol=7,header=False)


        self.trades_general_dict['rank'].to_excel(writer, sheet_name=f"{self._currency}_{self._money}",index=False,
                             startrow=5, startcol=1,header=False,na_rep='NA')

        startrow=17
        for trade in ['TX','DX','RX','IM']:
            self.trade_byproduct_table(writer, trade, startrow)
            startrow+=(self._toprank+3)

        self.part3_toexcel_ranking_cty(writer)
        self.adjust_excelformat_xlsxwriter(writer)

        writer.save()

        self.adjust_excelformat_openpyxl()

    def adjust_excelformat_xlsxwriter(self, writer):
        workbook  = writer.book
        worksheet = writer.sheets[f"{self._currency}_{self._money}"]
        title = f"HONG KONG'S TOP {self._toprank} TRADE WITH "+ self.countryname
        # setting format, labels, title, annotation for cells in the excel file
        merge_format_T = workbook.add_format({'bold': 1,'align': 'center',
                        'font_name': 'Arial','font_size':10})

        fmt_right = workbook.add_format({'align': 'right','font_name': "Arial",
                                'font_size':7.5})

        fmt_left = workbook.add_format({'align': 'left','font_name': "Arial",
                                    'font_size':7.5})

        fmt_bold = workbook.add_format({'bold': 1,'align': 'center','font_name': 'Arial',
                                        'font_size':7.5})

        fmt_bold_left = workbook.add_format({'bold': 1,'align': 'left','font_name': 'Arial',
                                            'font_size':7.5})

        wrap_text = workbook.add_format({'font_name': 'Arial','font_size':7.5,
                                        'text_wrap':True})

        worksheet.set_column('A:A', 22, None)
        worksheet.set_column('B:B', 45, wrap_text)
        worksheet.set_column('C:K', None, fmt_right)
        worksheet.write("A6", 'TOTAL EXPORTS', fmt_bold_left)
        worksheet.write("A7", 'DOMESTIC EXPORTS', fmt_bold_left)
        worksheet.write("A8", 'RE-EXPORTS', fmt_bold_left)
        worksheet.write("A9", 'IMPORTS', fmt_bold_left)
        worksheet.write("A10", '(OF WHICH RE-EXPORTED)', fmt_bold_left)
        worksheet.write("A11", 'TOTAL TRADE', fmt_bold_left)
        worksheet.write("A12", 'TRADE BALANCE', fmt_bold_left)

        worksheet.write(16, 1, '-TOTAL EXPORTS-', fmt_bold)
        worksheet.write(16+self._toprank+3, 1, '-DOMESTIC EXPORTS-', fmt_bold)
        worksheet.write(16+2*(self._toprank+3), 1, '-RE-EXPORTS-', fmt_bold)
        worksheet.write(16+3*(self._toprank+3), 1, '-IMPORTS-', fmt_bold)

        worksheet.merge_range('A1:J1', title, merge_format_T)
        worksheet.merge_range("H2:J2", f'VALUE : {self._currency} {self._money}', fmt_bold)
        worksheet.merge_range("H3:J3", '% CHANGE', fmt_bold)
        worksheet.merge_range("A14:B14", 'MAJOR COMMODITIES OF TRADES', fmt_bold)
        worksheet.write("A16", "SITC", fmt_bold)

        # writing VALUE and % SHARE labels
        for c in range(2,10,2):
            worksheet.write(15, c, "VALUE", fmt_bold)
            worksheet.write(15, c+1, "% SHARE", fmt_bold)
        worksheet.write(15, 10, "% CHG", fmt_bold)

        # year
        unique_yr = sorted(set([str(yr)[:4]for yr in self._periods]))
        for i, yr in enumerate(unique_yr):
            worksheet.write(4, 3+i, int(yr), fmt_bold)
        worksheet.merge_range("C15:D15", int(unique_yr[0]), fmt_bold)
        worksheet.merge_range("E15:F15", int(unique_yr[1]), fmt_bold)
        worksheet.merge_range("G15:H15", int(unique_yr[2]), fmt_bold)
        worksheet.merge_range("I15:K15", int(unique_yr[3]), fmt_bold)

        worksheet.write("H5", f"'{unique_yr[1][-2:]}/{unique_yr[0][-2:]}", fmt_bold)
        worksheet.write("I5", f"'{unique_yr[2][-2:]}/{unique_yr[1][-2:]}", fmt_bold)
        worksheet.write("J5", f"'{unique_yr[3][-2:]}/{unique_yr[2][-2:]}", fmt_bold)

        # acquire latest month name
        month_name = calendar.month_abbr[int(str(self._lastperiod)[-2:])]

        worksheet.write("G4", "JAN-"+month_name.upper(), fmt_bold)
        worksheet.write("J4", "JAN-"+month_name.upper(), fmt_bold)
        worksheet.merge_range("I14:K14", "JAN-"+month_name.upper(), fmt_bold)

        # hide gridlines
        worksheet.hide_gridlines(2)

        # write source, symbol annotation at the end
        worksheet.merge_range(18+4*(self._toprank+3),0,18+4*(self._toprank+3),5, "* INSIGNIFICANT            ∞ INFINITY", fmt_left)
        worksheet.merge_range(19+4*(self._toprank+3),0,19+4*(self._toprank+3),5, "..OVER 1000% INCREASE      - NIL     N.E.S. NOT ELSEWHERE SPECIFIED", fmt_left)
        worksheet.merge_range(20+4*(self._toprank+3),0,20+4*(self._toprank+3),5, "SOURCE: HONG KONG TRADE STATISTICS, CENSUS & STATISTICS DEPT.", fmt_left)

    def trade_byproduct_table(self, writer, trade, startrow=17):
        self.trades_byproduct_dict[trade]['figures'].index.to_frame(index=False).to_excel(writer,sheet_name=f"{self._currency}_{self._money}",index=False,startrow=startrow, startcol=0,header=False)

        for i, c in zip([0,1,2,3],[2,4,6,8]):
            ExcelOutput.denotesymbol(self.trades_byproduct_dict[trade]['figures'].iloc[:,i], "figures").to_excel(writer,sheet_name=f"{self._currency}_{self._money}",index=False,startrow=startrow, startcol=c,header=False)
            ExcelOutput.denotesymbol(self.trades_byproduct_dict[trade]['percent_share'].iloc[:,i], "share").to_excel(writer,sheet_name=f"{self._currency}_{self._money}",index=False,startrow=startrow, startcol=c+1,header=False)

        ExcelOutput.denotesymbol(self.trades_byproduct_dict[trade]['percent_change'].iloc[:,-1], "percent_change").to_excel(writer,sheet_name=f"{self._currency}_{self._money}",index=False,startrow=startrow, startcol=10,header=False)

    def part3_toexcel_ranking_cty(self, writer):
        workbook  = writer.book
        worksheet = writer.sheets[f"{self._currency}_{self._money}"]
        fmt_bold = workbook.add_format({'bold': 1,'align': 'center','font_name': 'Arial',
                                            'font_size':7.5})
        fmt_bold_right = workbook.add_format({'bold': 1,'align': 'right','font_name': 'Arial',
                                            'font_size':7.5})
        #acquire latest month name
        month_name = calendar.month_abbr[int(str(self._periods[-1])[-2:])]

        worksheet.write("C3", 'RANKING', fmt_bold)
        worksheet.write("C4", "JAN-"+month_name.upper(), fmt_bold)
        worksheet.write("B5", "'"+str(self._periods[-2])[2:4], fmt_bold_right)
        worksheet.write("C5", "'"+str(self._periods[-1])[2:4], fmt_bold_right)

    def adjust_excelformat_openpyxl(self):
        wb = openpyxl.load_workbook(f"{self._file_path}/{self._excel_name}")
        ws = wb[f"{self._currency}_{self._money}"]

        # use openpyxl
        ft1 = openpyxl.styles.Font(name='Arial', size=7.5)
        value_format = '#,##0'
        pct_format = '#,##0.0'

        # no border format
        side = openpyxl.styles.Side(border_style=None)
        no_border = openpyxl.styles.borders.Border(
            left=side,
            right=side,
            top=side,
            bottom=side,
        )
        # row and column number start from 1, not 0 index
        ##### set upper part value format
        # set D6:G12 value format
        for row in range(6,13):
            for col in range(4,8):
                _cell = ws.cell(row,col)
                _cell.number_format = value_format
                _cell.font = ft1

        # set H6:J11 %CHG format
        for row in range(6,13):
            for col in range(8,11):
                _cell = ws.cell(row,col)
                _cell.font = ft1

        # set A:A SITC code format, up to row no.1000
        for row in range(18,1000):
            _cell = ws.cell(row,1)
            _cell.border = no_border
            _cell.font = ft1
            _cell.alignment = openpyxl.styles.Alignment(horizontal='left')
            _cell.number_format = '@'

        # set C:I value format, up to row no.1000
        for row in range(18,1000):
            for col in range(3,10,2):
                _cell = ws.cell(row,col)
                _cell.number_format = value_format
                _cell.font = ft1

        # set D:K %CHG format, up to row no.1000
        for row in range(18,1000):
            for col in [4,6,8,10,11]:
                _cell = ws.cell(row,col)
                _cell.font = ft1
                _cell.number_format = pct_format

        # set upper trade fig %CHG format
        for row in range(6,12):
            for col in range(8,11):
                _cell = ws.cell(row,col)
                _cell.font = ft1
                _cell.number_format = pct_format
        wb.save(f"{self._file_path}/{self._excel_name}")

def autofit(excel_name, currency, money):
    """using win32com.client to control excel to autofit"""
    try:
        excel = Dispatch('Excel.Application')
    except e as exception:
        print(e)
        print("Error exists while win32 connects with Excel Application")

    thisdir = os.getcwd()
    wb = excel.Workbooks.Open(thisdir+"/"+excel_name)
    ws = wb.Worksheets(f"{currency}_{money}")
    ws.Columns.AutoFit()
    wb.Save()
    wb.Close()
